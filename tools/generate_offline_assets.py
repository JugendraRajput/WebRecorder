#!/usr/bin/env python3
"""
Generate WebRecorder asset bundles from Excel sheets.

Output format:
  <output_dir>/offline_pages/<folder_name>/
    - <md5(url)>.html (default) or .mht (optional)
    - metadata.json

Dependencies:
  pip install openpyxl requests beautifulsoup4
"""

from __future__ import annotations

import argparse
import hashlib
import json
import mimetypes
import re
import sys
import time
from dataclasses import dataclass
from email.utils import formatdate
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 " \
             "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"


@dataclass
class RowItem:
    row_index: int
    title: str
    url: str


def sanitize_folder_name(raw_name: str) -> str:
    sanitized = re.sub(r"\s+", "_", raw_name.strip())
    sanitized = re.sub(r"[^A-Za-z0-9_-]", "_", sanitized)
    sanitized = re.sub(r"_+", "_", sanitized)
    if not sanitized:
        sanitized = "default_list"
    return sanitized[:80]


def build_offline_file_name(url: str, output_format: str) -> str:
    ext = ".mht" if output_format == "mht" else ".html"
    return hashlib.md5(url.encode("utf-8")).hexdigest() + ext


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


UNWANTED_KEYWORDS = {"sex", "porn", "xxx", "adult", "nude", "abuse", "erotic", "hentai"}


def is_unwanted_title(title: str) -> bool:
    if not title:
        return False
    lower_title = title.lower()
    words = re.findall(r"\b\w+\b", lower_title)
    for kw in UNWANTED_KEYWORDS:
        if kw in words or kw in lower_title:
            return True
    return False


def parse_excel_rows(excel_path: Path) -> List[RowItem]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: List[RowItem] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
        title = normalize_cell(row[0])
        url = normalize_cell(row[1])
        if not title or not url:
            continue
        if is_unwanted_title(title):
            print(f"  [FILTERED] Skipping unwanted/abusive title: '{title}'")
            continue
        rows.append(RowItem(row_index=row_index, title=title, url=url))

    workbook.close()
    return rows


def find_excel_files(folder_path: Path, recursive: bool = True) -> List[Path]:
    iterator = folder_path.rglob("*.xlsx") if recursive else folder_path.glob("*.xlsx")
    files = [p for p in iterator if p.is_file() and not p.name.startswith("~$")]
    files.sort(key=lambda p: str(p).lower())
    return files


def pick_src_from_srcset(srcset: str) -> Optional[str]:
    if not srcset:
        return None
    first = srcset.split(",")[0].strip()
    if not first:
        return None
    return first.split(" ")[0].strip() or None


def resolve_mime_type(url: str, response_content_type: str) -> str:
    if response_content_type:
        clean = response_content_type.split(";", 1)[0].strip().lower()
        if clean.startswith("image/"):
            return clean
    guessed, _ = mimetypes.guess_type(url)
    return guessed or "image/png"


def resolve_resource_mime_type(url: str, response_content_type: str) -> str:
    if response_content_type:
        clean = response_content_type.split(";", 1)[0].strip().lower()
        if clean:
            return clean
    guessed, _ = mimetypes.guess_type(url)
    return guessed or "application/octet-stream"


def is_non_downloadable_url(candidate: str) -> bool:
    lower = candidate.lower()
    return (
        not candidate
        or lower.startswith("data:")
        or lower.startswith("blob:")
        or lower.startswith("javascript:")
        or lower.startswith("#")
    )


def download_binary_as_data_uri(
    session: requests.Session,
    resource_url: str,
    referer_url: str,
    timeout_seconds: int,
    max_bytes: int,
    accept_header: str,
) -> Optional[str]:
    request_headers = {
        "Referer": referer_url,
        "Accept": accept_header,
        "Accept-Language": "en-US,en;q=0.9",
    }

    for headers in (request_headers, {"Accept": accept_header}):
        try:
            response = session.get(
                resource_url,
                timeout=timeout_seconds,
                allow_redirects=True,
                headers=headers,
            )
            response.raise_for_status()
            content = response.content
            if not content or len(content) > max_bytes:
                return None

            mime_type = resolve_resource_mime_type(resource_url, response.headers.get("Content-Type", ""))
            encoded = __import__("base64").b64encode(content).decode("ascii")
            return f"data:{mime_type};base64,{encoded}"
        except Exception:
            continue

    return None


def download_image_as_data_uri(
    session: requests.Session,
    image_url: str,
    referer_url: str,
    timeout_seconds: int,
    max_image_bytes: int,
) -> Optional[str]:
    return download_binary_as_data_uri(
        session=session,
        resource_url=image_url,
        referer_url=referer_url,
        timeout_seconds=timeout_seconds,
        max_bytes=max_image_bytes,
        accept_header="image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
    )


def rewrite_css_urls_to_absolute(css_text: str, css_url: str) -> str:
    pattern = re.compile(r"url\(([^)]+)\)", re.IGNORECASE)

    def replace_url(match: re.Match[str]) -> str:
        original = match.group(1).strip()
        if not original:
            return match.group(0)

        cleaned = normalize_asset_url(original.strip("\"'").strip())
        if is_non_downloadable_url(cleaned):
            return match.group(0)
        if cleaned.startswith("http://") or cleaned.startswith("https://"):
            return f'url("{cleaned}")'

        absolute = urljoin(css_url, cleaned)
        return f'url("{absolute}")'

    return pattern.sub(replace_url, css_text)


def inline_css_url_resources(
    session: requests.Session,
    css_text: str,
    css_base_url: str,
    referer_url: str,
    timeout_seconds: int,
    max_asset_bytes: int,
    cache: Optional[Dict[str, Optional[str]]] = None,
) -> Tuple[str, int]:
    if not css_text:
        return css_text, 0

    if cache is None:
        cache = {}

    replaced_count = 0
    pattern = re.compile(r"url\(([^)]+)\)", re.IGNORECASE)

    def replace_url(match: re.Match[str]) -> str:
        nonlocal replaced_count
        original = match.group(1).strip()
        if not original:
            return match.group(0)

        cleaned = normalize_asset_url(original.strip("\"'").strip())
        if is_non_downloadable_url(cleaned):
            return match.group(0)

        absolute = urljoin(css_base_url, cleaned)
        if absolute not in cache:
            cache[absolute] = download_binary_as_data_uri(
                session=session,
                resource_url=absolute,
                referer_url=referer_url,
                timeout_seconds=timeout_seconds,
                max_bytes=max_asset_bytes,
                accept_header="*/*",
            )

        data_uri = cache.get(absolute)
        if data_uri:
            replaced_count += 1
            return f'url("{data_uri}")'

        return f'url("{absolute}")'

    return pattern.sub(replace_url, css_text), replaced_count


def inline_stylesheets_in_html(
    session: requests.Session,
    html: str,
    base_url: str,
    timeout_seconds: int,
    max_asset_bytes: int,
) -> str:
    soup = BeautifulSoup(html, "html.parser")
    resource_cache: Dict[str, Optional[str]] = {}

    for link in soup.find_all("link"):
        rel = link.get("rel") or []
        if isinstance(rel, str):
            rel_tokens = [rel.lower()]
        else:
            rel_tokens = [str(token).lower() for token in rel]

        if "stylesheet" not in rel_tokens:
            continue

        href = (link.get("href") or "").strip()
        if not href:
            continue

        css_url = urljoin(base_url, href)
        css_text = fetch_text_resource(session, css_url, timeout_seconds)
        if not css_text:
            continue

        css_text = rewrite_css_urls_to_absolute(css_text, css_url)
        css_text, _ = inline_css_url_resources(
            session=session,
            css_text=css_text,
            css_base_url=css_url,
            referer_url=base_url,
            timeout_seconds=timeout_seconds,
            max_asset_bytes=max_asset_bytes,
            cache=resource_cache,
        )
        style_tag = soup.new_tag("style")
        style_tag.string = css_text
        link.replace_with(style_tag)

    for style_tag in soup.find_all("style"):
        style_text = style_tag.string if style_tag.string is not None else style_tag.get_text()
        inlined_css, _ = inline_css_url_resources(
            session=session,
            css_text=style_text,
            css_base_url=base_url,
            referer_url=base_url,
            timeout_seconds=timeout_seconds,
            max_asset_bytes=max_asset_bytes,
            cache=resource_cache,
        )
        style_tag.string = inlined_css

    for node in soup.find_all(style=True):
        style_value = node.get("style", "")
        inlined_style, _ = inline_css_url_resources(
            session=session,
            css_text=style_value,
            css_base_url=base_url,
            referer_url=base_url,
            timeout_seconds=timeout_seconds,
            max_asset_bytes=max_asset_bytes,
            cache=resource_cache,
        )
        node["style"] = inlined_style

    return str(soup)


def normalize_asset_url(raw_url: str) -> str:
    cleaned = (raw_url or "").strip()
    if cleaned.startswith("//"):
        return "https:" + cleaned
    return cleaned


def extract_image_candidate(tag) -> Optional[str]:
    direct_attrs = [
        "src",
        "data-src",
        "data-original",
        "data-lazy-src",
        "data-fallback-src",
        "data-src-retina",
    ]
    srcset_attrs = ["srcset", "data-srcset", "data-lazy-srcset"]

    for attr in direct_attrs:
        value = normalize_asset_url(tag.get(attr) or "")
        if value and not value.startswith("data:"):
            return value

    for attr in srcset_attrs:
        value = normalize_asset_url(tag.get(attr) or "")
        chosen = pick_src_from_srcset(value)
        if chosen and not chosen.startswith("data:"):
            return chosen

    return None


def apply_data_uri_to_tag(tag, data_uri: str) -> None:
    if tag.name == "source":
        tag["srcset"] = data_uri
        for attr in ("src", "data-src", "data-srcset", "data-lazy-srcset"):
            if attr in tag.attrs:
                del tag.attrs[attr]
        return

    tag["src"] = data_uri
    for attr in (
        "srcset",
        "data-src",
        "data-original",
        "data-lazy-src",
        "data-fallback-src",
        "data-src-retina",
        "data-srcset",
        "data-lazy-srcset",
    ):
        if attr in tag.attrs:
            del tag.attrs[attr]


def inline_images_in_html(
    session: requests.Session,
    html: str,
    base_url: str,
    timeout_seconds: int,
    max_image_bytes: int,
    max_total_bytes: int,
) -> str:
    soup = BeautifulSoup(html, "html.parser")

    image_nodes = list(soup.find_all("img")) + list(soup.find_all("source"))
    image_url_map: Dict[str, str] = {}
    total_bytes = 0

    for node in image_nodes:
        candidate = extract_image_candidate(node)
        if not candidate:
            continue

        absolute_url = urljoin(base_url, candidate)
        if absolute_url in image_url_map:
            continue

        data_uri = download_image_as_data_uri(
            session=session,
            image_url=absolute_url,
            referer_url=base_url,
            timeout_seconds=timeout_seconds,
            max_image_bytes=max_image_bytes,
        )
        if not data_uri:
            continue

        estimated_bytes = len(data_uri.encode("utf-8"))
        if total_bytes + estimated_bytes > max_total_bytes:
            break

        total_bytes += estimated_bytes
        image_url_map[absolute_url] = data_uri

    if not image_url_map:
        return str(soup)

    for node in image_nodes:
        candidate = extract_image_candidate(node)
        if not candidate:
            continue

        absolute_url = urljoin(base_url, candidate)
        data_uri = image_url_map.get(absolute_url)
        if not data_uri:
            continue

        apply_data_uri_to_tag(node, data_uri)

    return str(soup)


def collect_missing_image_urls(html: str, base_url: str) -> List[str]:
    soup = BeautifulSoup(html, "html.parser")
    missing: Set[str] = set()

    for node in list(soup.find_all("img")) + list(soup.find_all("source")):
        candidate = extract_image_candidate(node)
        if not candidate or candidate.startswith("data:"):
            continue
        missing.add(urljoin(base_url, candidate))

    return sorted(missing)


def collect_missing_css_resource_urls(html: str, base_url: str) -> List[str]:
    soup = BeautifulSoup(html, "html.parser")
    missing: Set[str] = set()
    pattern = re.compile(r"url\(([^)]+)\)", re.IGNORECASE)

    def collect_from_css(css_text: str, css_base_url: str) -> None:
        for match in pattern.finditer(css_text or ""):
            raw = match.group(1).strip()
            cleaned = normalize_asset_url(raw.strip("\"'").strip())
            if is_non_downloadable_url(cleaned):
                continue
            missing.add(urljoin(css_base_url, cleaned))

    for style_tag in soup.find_all("style"):
        style_text = style_tag.string if style_tag.string is not None else style_tag.get_text()
        collect_from_css(style_text or "", base_url)

    for node in soup.find_all(style=True):
        collect_from_css(node.get("style", ""), base_url)

    return sorted(missing)


def retry_missing_css_resources_in_html(
    session: requests.Session,
    html: str,
    base_url: str,
    timeout_seconds: int,
    max_asset_bytes: int,
) -> Tuple[str, int]:
    soup = BeautifulSoup(html, "html.parser")
    replaced_total = 0
    cache: Dict[str, Optional[str]] = {}

    for style_tag in soup.find_all("style"):
        style_text = style_tag.string if style_tag.string is not None else style_tag.get_text()
        inlined_css, replaced_count = inline_css_url_resources(
            session=session,
            css_text=style_text or "",
            css_base_url=base_url,
            referer_url=base_url,
            timeout_seconds=timeout_seconds,
            max_asset_bytes=max_asset_bytes,
            cache=cache,
        )
        if replaced_count > 0:
            style_tag.string = inlined_css
            replaced_total += replaced_count

    for node in soup.find_all(style=True):
        inlined_style, replaced_count = inline_css_url_resources(
            session=session,
            css_text=node.get("style", ""),
            css_base_url=base_url,
            referer_url=base_url,
            timeout_seconds=timeout_seconds,
            max_asset_bytes=max_asset_bytes,
            cache=cache,
        )
        if replaced_count > 0:
            node["style"] = inlined_style
            replaced_total += replaced_count

    return str(soup), replaced_total


def retry_missing_images_in_html(
    session: requests.Session,
    html: str,
    base_url: str,
    timeout_seconds: int,
    max_image_bytes: int,
    max_total_bytes: int,
) -> Tuple[str, int]:
    soup = BeautifulSoup(html, "html.parser")
    fixed_count = 0
    total_bytes = 0

    for node in list(soup.find_all("img")) + list(soup.find_all("source")):
        candidate = extract_image_candidate(node)
        if not candidate or candidate.startswith("data:"):
            continue

        absolute_url = urljoin(base_url, candidate)
        data_uri = download_image_as_data_uri(
            session=session,
            image_url=absolute_url,
            referer_url=base_url,
            timeout_seconds=timeout_seconds,
            max_image_bytes=max_image_bytes,
        )
        if not data_uri:
            continue

        estimated_bytes = len(data_uri.encode("utf-8"))
        if total_bytes + estimated_bytes > max_total_bytes:
            break

        total_bytes += estimated_bytes
        apply_data_uri_to_tag(node, data_uri)
        fixed_count += 1

    return str(soup), fixed_count


def fetch_text_resource(session: requests.Session, url: str, timeout_seconds: int) -> Optional[str]:
    try:
        response = session.get(url, timeout=timeout_seconds, allow_redirects=True)
        response.raise_for_status()
        if not response.encoding:
            response.encoding = response.apparent_encoding
        return response.text
    except Exception:
        return None


def fetch_html(session: requests.Session, url: str, timeout_seconds: int) -> Optional[str]:
    return fetch_text_resource(session, url, timeout_seconds)


def build_mht_archive(html: str, source_url: str, title: str) -> bytes:
    boundary = "----=_NextPart_" + hashlib.md5((source_url + title).encode("utf-8", errors="ignore")).hexdigest()[:16]
    safe_title = (title or "WebRecorder Offline Page").encode("ascii", errors="ignore").decode("ascii").strip()
    if not safe_title:
        safe_title = "WebRecorder Offline Page"

    header_lines = [
        "From: <Saved by WebRecorder>",
        f"Subject: {safe_title}",
        f"Date: {formatdate(localtime=True)}",
        "MIME-Version: 1.0",
        f"Content-Type: multipart/related; type=\"text/html\"; boundary=\"{boundary}\"",
        "",
        f"--{boundary}",
        "Content-Type: text/html; charset=\"utf-8\"",
        "Content-Transfer-Encoding: 8bit",
        f"Content-Location: {source_url}",
        "",
    ]

    header_blob = "\r\n".join(header_lines).encode("utf-8", errors="ignore")
    html_blob = html.encode("utf-8", errors="ignore")
    footer_blob = f"\r\n--{boundary}--\r\n".encode("utf-8", errors="ignore")
    return header_blob + html_blob + footer_blob


def save_html_with_recheck(
    session: requests.Session,
    output_file: Path,
    html: str,
    base_url: str,
    timeout_seconds: int,
    max_image_bytes: int,
    max_total_bytes: int,
    max_css_asset_bytes: int,
    recheck_attempts: int,
) -> None:
    output_file.write_text(html, encoding="utf-8", errors="ignore")

    for attempt in range(1, recheck_attempts + 1):
        saved_html = output_file.read_text(encoding="utf-8", errors="ignore")
        missing_images = collect_missing_image_urls(saved_html, base_url)
        missing_css = collect_missing_css_resource_urls(saved_html, base_url)
        if not missing_images and not missing_css:
            return

        print(
            f"    Recheck {attempt}/{recheck_attempts}: "
            f"{len(missing_images)} image(s), {len(missing_css)} css-resource(s) still external, retrying..."
        )

        repaired_html, fixed_images = retry_missing_images_in_html(
            session=session,
            html=saved_html,
            base_url=base_url,
            timeout_seconds=timeout_seconds,
            max_image_bytes=max_image_bytes,
            max_total_bytes=max_total_bytes,
        )

        repaired_html, fixed_css = retry_missing_css_resources_in_html(
            session=session,
            html=repaired_html,
            base_url=base_url,
            timeout_seconds=timeout_seconds,
            max_asset_bytes=max_css_asset_bytes,
        )

        if fixed_images + fixed_css == 0:
            return

        output_file.write_text(repaired_html, encoding="utf-8", errors="ignore")


def generate_bundle(
    excel_path: Path,
    output_dir: Path,
    folder_name: str,
    timeout_seconds: int,
    max_image_mb: int,
    max_total_mb: int,
    max_css_asset_mb: int,
    output_format: str,
    recheck_attempts: int,
) -> Tuple[int, int, Path]:
    rows = parse_excel_rows(excel_path)
    if not rows:
        raise RuntimeError("No valid rows found in the Excel sheet (expected title in column A and URL in column B).")

    bundle_dir = output_dir / "offline_pages" / folder_name
    bundle_dir.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    metadata: Dict[str, dict] = {}
    success_count = 0

    max_image_bytes = max_image_mb * 1024 * 1024
    max_total_bytes = max_total_mb * 1024 * 1024
    max_css_asset_bytes = max_css_asset_mb * 1024 * 1024

    for idx, item in enumerate(rows, start=1):
        print(f"  [{idx}/{len(rows)}] Fetching: {item.title}")
        html = fetch_html(session, item.url, timeout_seconds)
        if not html:
            print(f"    ! Failed: {item.url}")
            continue

        styled_html = inline_stylesheets_in_html(
            session=session,
            html=html,
            base_url=item.url,
            timeout_seconds=timeout_seconds,
            max_asset_bytes=max_css_asset_bytes,
        )

        offline_html = inline_images_in_html(
            session=session,
            html=styled_html,
            base_url=item.url,
            timeout_seconds=timeout_seconds,
            max_image_bytes=max_image_bytes,
            max_total_bytes=max_total_bytes,
        )

        file_name = build_offline_file_name(item.url, output_format)
        output_file = bundle_dir / file_name
        if output_format == "mht":
            output_file.write_bytes(build_mht_archive(offline_html, item.url, item.title))
        else:
            save_html_with_recheck(
                session=session,
                output_file=output_file,
                html=offline_html,
                base_url=item.url,
                timeout_seconds=timeout_seconds,
                max_image_bytes=max_image_bytes,
                max_total_bytes=max_total_bytes,
                max_css_asset_bytes=max_css_asset_bytes,
                recheck_attempts=recheck_attempts,
            )

        metadata[file_name] = {
            "original_url": item.url,
            "title": item.title,
            "row_index": item.row_index,
            "updated_at": int(time.time()),
        }
        success_count += 1

    (bundle_dir / "metadata.json").write_text(
        json.dumps(metadata, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    return success_count, len(rows), bundle_dir


def ask(prompt: str, default: Optional[str] = None) -> str:
    suffix = f" [{default}]" if default else ""
    value = input(f"{prompt}{suffix}: ").strip()
    if not value and default is not None:
        return default
    return value


def resolve_excel_inputs(args) -> List[Path]:
    if args.excel:
        excel_path = Path(args.excel.strip('"')).expanduser().resolve()
        if not excel_path.exists() or excel_path.suffix.lower() != ".xlsx":
            raise RuntimeError("Excel file not found or unsupported format. Use .xlsx.")
        return [excel_path]

    if args.excel_folder:
        folder_path = Path(args.excel_folder.strip('"')).expanduser().resolve()
    else:
        source_input = ask("Excel file path (.xlsx) or folder path containing Excel files")
        folder_or_file = Path(source_input.strip('"')).expanduser().resolve()
        if folder_or_file.is_file():
            if folder_or_file.suffix.lower() != ".xlsx":
                raise RuntimeError("File is not .xlsx")
            return [folder_or_file]
        folder_path = folder_or_file

    if not folder_path.exists() or not folder_path.is_dir():
        raise RuntimeError("Excel folder not found.")

    excel_files = find_excel_files(folder_path, recursive=not args.no_recursive)
    if not excel_files:
        raise RuntimeError("No .xlsx files found in folder.")

def upload_file_to_server(server_base_url: str, folder_name: str, file_path: Path, title: str, original_url: str, row_index: int) -> bool:
    endpoint = f"{server_base_url.rstrip('/')}/sync_api.php"
    data = {
        "action": "upload",
        "folder": folder_name,
        "filename": file_path.name,
        "title": title,
        "original_url": original_url,
        "row_index": str(row_index),
    }
    with open(file_path, "rb") as f:
        files = {"html_file": (file_path.name, f, "text/html")}
        try:
            res = requests.post(endpoint, data=data, files=files, timeout=30)
            return res.status_code == 200 and res.json().get("ok", False)
        except Exception as e:
            print(f"    ! Upload error: {e}")
            return False


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate WebRecorder offline asset bundle from Excel.")
    parser.add_argument("--excel", help="Path to one .xlsx Excel file")
    parser.add_argument("--excel-folder", help="Folder path containing .xlsx files")
    parser.add_argument("--no-recursive", action="store_true", help="Only scan top-level folder for .xlsx files")
    parser.add_argument("--output", help="Output directory (default: current directory)")
    parser.add_argument("--folder", help="Asset folder name (used only for single Excel input)")
    parser.add_argument("--format", choices=["html", "mht"], default="html", help="Offline file format (default: html)")
    parser.add_argument("--timeout", type=int, default=25, help="HTTP timeout seconds")
    parser.add_argument("--max-image-mb", type=int, default=16, help="Max bytes per image in MB")
    parser.add_argument("--max-total-mb", type=int, default=120, help="Max total embedded image data per page in MB")
    parser.add_argument("--max-css-asset-mb", type=int, default=10, help="Max bytes per CSS resource (background image/font) in MB")
    parser.add_argument("--recheck-attempts", type=int, default=2, help="Retry passes for missing image/CSS resources after saving each page")
    parser.add_argument("--upload-server", help="Server base URL to upload generated assets to (e.g. https://webrecorder.jdworks.in)")
    args = parser.parse_args()

    try:
        excel_files = resolve_excel_inputs(args)
    except Exception as exc:
        print(f"Input error: {exc}")
        return 1

    output_input = args.output or ask("Output folder", str(Path.cwd()))
    output_dir = Path(output_input.strip('"')).expanduser().resolve()

    if args.format == "mht":
        print("Warning: externally-generated MHT can be inconsistent in Android WebView.")
        print("If pages appear blank, regenerate with --format html.\n")

    used_folder_names: Set[str] = set()
    grand_success = 0
    grand_total = 0

    print("\nGenerating offline bundle(s)...")
    print(f"Excel files found: {len(excel_files)}")
    print(f"Output:            {output_dir}")
    print(f"Format:            {args.format}")
    print(f"Image limit/page:  {args.max_total_mb} MB")
    print(f"Image limit/file:  {args.max_image_mb} MB")
    print(f"CSS asset limit:  {args.max_css_asset_mb} MB")
    print(f"Recheck attempts:  {args.recheck_attempts}\n")

    for index, excel_path in enumerate(excel_files, start=1):
        if len(excel_files) == 1 and args.folder:
            folder_name = sanitize_folder_name(args.folder)
        else:
            folder_name = sanitize_folder_name(excel_path.stem)

        base_name = folder_name
        suffix = 2
        while folder_name in used_folder_names:
            folder_name = f"{base_name}_{suffix}"
            suffix += 1
        used_folder_names.add(folder_name)

        print(f"[{index}/{len(excel_files)}] Excel: {excel_path.name} -> folder: {folder_name}")
        try:
            success_count, total_rows, bundle_dir = generate_bundle(
                excel_path=excel_path,
                output_dir=output_dir,
                folder_name=folder_name,
                timeout_seconds=args.timeout,
                max_image_mb=args.max_image_mb,
                max_total_mb=args.max_total_mb,
                max_css_asset_mb=args.max_css_asset_mb,
                output_format=args.format,
                recheck_attempts=args.recheck_attempts,
            )
            grand_success += success_count
            grand_total += total_rows
            print(f"  Done: {success_count}/{total_rows} -> {bundle_dir}\n")

            if args.upload_server:
                print(f"  Uploading folder '{folder_name}' to server: {args.upload_server} ...")
                metadata_file = bundle_dir / "metadata.json"
                if metadata_file.exists():
                    meta_data = json.loads(metadata_file.read_text(encoding="utf-8"))
                    upload_success = 0
                    for fname, entry in meta_data.items():
                        fpath = bundle_dir / fname
                        if fpath.exists():
                            ok = upload_file_to_server(
                                server_base_url=args.upload_server,
                                folder_name=folder_name,
                                file_path=fpath,
                                title=entry.get("title", ""),
                                original_url=entry.get("original_url", ""),
                                row_index=entry.get("row_index", -1),
                            )
                            if ok:
                                upload_success += 1
                    print(f"  Uploaded: {upload_success}/{len(meta_data)} files to server.\n")

        except Exception as exc:
            print(f"  Failed: {exc}\n")

    print("All done.")
    print(f"Saved pages total: {grand_success}/{grand_total}")
    print("Copy generated folder(s) into Android assets path:")
    print("  app/src/main/assets/offline_pages/<your_folder>")
    print("Then use app option: Select Source -> Load From Assets")
    return 0


if __name__ == "__main__":
    sys.exit(main())



